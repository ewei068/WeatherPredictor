import io
import os
import random
import praw
import random
import requests
import xlwt
from xlwt import Workbook
import openpyxl
import pandas as pd
import csv


#GOOGLE_APPLICATION_CREDENTIALS='/Users/elviswei/PycharmProjects/CalHacks/apikey.json'
'''
# Instantiates a client
client = vision.ImageAnnotatorClient()

# The name of the image file to annotate
file_name = os.path.abspath('pic.jpg')

# Loads the image into memory
with io.open(file_name, 'rb') as image_file:
    content = image_file.read()

image = types.Image(content=content)

# Performs label detection on the image file
response = client.label_detection(image=image)
labels = response.label_annotations

print(response)
'''
'''
reddit = praw.Reddit(client_id='BFWDzf7vpDnkiA',
                     client_secret='Bq9Kn4yB0rbJGdQZAwOEJoIb9x8',
                     user_agent='CalHacks Project by u/Pokestar9999')
memes_submissions = reddit.subreddit('okbuddyretard').top('week')
post_to_pick = random.randint(1, 12)
for i in range(0, post_to_pick):
    submission = next(x for x in memes_submissions if not x.stickied)
'''
source = "10yr.csv"
dest = "10yrSOD.csv"

source_df = pd.read_csv(source)
rows = len(source_df)
source_df = source_df.T
 # thing = source_df[300]
# print(source_df)
dest_rows = []
column_names = ['DATE', "DailyAverageDryBulbTemperature", "DailyAverageRelativeHumidity", "DailyAverageStationPressure", "DailyAverageWindSpeed", "DailyDepartureFromNormalAverageTemperature",
                "DailyMaximumDryBulbTemperature", "DailyMinimumDryBulbTemperature", "DailyPrecipitation", "DailySnowDepth", "DailySnowfall", "DailyWeather"]
dest_rows.append(column_names)
for i in range(1, rows):
    # print(source_df[i]["REPORT_TYPE"])
    if source_df[i]['REPORT_TYPE'] == "SOD  ":
        new_row = [source_df[i][column_name] for column_name in column_names]
        dest_rows.append(new_row)
# print(dest_rows)

with open(dest, "w", newline="") as file:
    writer = csv.writer(file, quoting=csv.QUOTE_ALL)
    for dest_row in dest_rows:
        writer.writerow(dest_row)

'''
filename = "testing.xlsx"

wb = openpyxl.load_workbook(filename=filename)
ws = wb['Sheet1']

iterations = 0

while iterations < 10:
    start = random.choice(range(1514764800, 1535760000))
    end = start + 3600
    api_url = 'https://api.pushshift.io/reddit/submission/search/?subreddit=pics&after=' + str(start) + '&before=' + str(end)
    response = requests.get(api_url)
    data = response.json()

    for post in data["data"]:
        new_row = ws.max_row + 1
        ws.cell(row=new_row, column=1, value=post["full_link"])
        ws.cell(row=new_row, column=2, value=post["url"])
        ws.cell(row=new_row, column=3, value=post["title"])
        ws.cell(row=new_row, column=4, value=post["created_utc"])
        ws.cell(row=new_row, column=5, value=post["score"])
        ws.cell(row=new_row, column=6, value=post["num_comments"])

    print('Iteration', iterations)

    if iterations % 25 == 0:
        wb.save(filename)
        print('Save Complete')
        print('Row:', ws.max_row)

    if iterations % 100 == 0:
        print('Backup Started')
        wb.save('backup.xlsx')
        print('Backup Complete')

    iterations += 1

wb.save(filename)
'''

# Writing to an excel
# sheet using Pyt

# Workbook is created
